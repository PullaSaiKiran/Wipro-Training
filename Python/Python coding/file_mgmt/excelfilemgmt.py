import os
from openpyxl import Workbook, load_workbook

def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Age"])
    sheet.append(["Jon Doe", 30])
    sheet.append(["Jane Smith", 25])
    workbook.save(filename)
    print(f"Wrote {filename} successfully")

def read_excel(filename):
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
        # Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, age = row
            print(f"Name: {name}, Age: {age}")
    else:
        print(f"{filename} does not exist.")

def delete_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print(f"Deleted {filename} successfully")
    else:
        print(f"{filename} not found")

# Example usage
filename = "data.xlsx"
write_excel(filename)
print("Data read from excel file:")
read_excel(filename)
# delete_excel(filename)
